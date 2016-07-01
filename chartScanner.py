#import library
import openpyxl
#prompt for excel file and open it
exfile = input("Enter the excel file you wish to open: ")
wb = openpyxl.load_workbook(exfile)
wb.get_sheet_names()
#prompt for sheet
page = input("Which sheet do you wish to work with: ")
sheet = wb.get_sheet_by_name(page)
#save value for highest row and column, will be used later for loops
words = input("What words do you want to search for: ")
words.split()
maxRow = sheet.max_row
#print ("Max Row: ", maxRow)
maxCol = sheet.max_column
#print ("Max Column: ", maxCol)

#flag will let you break out of the loop if a word is ever met
flag = false
for i in range (2,maxRow):
	for word in words:
		if word in sheet.cell(row=i, column = 2).value:
			#print (i);
			print ("{0} found in row {1} column {2}", word, i , 2)
			flag = true
			break
	if flag = true:
		flag = false
		continue
	print ("Look at row {0}",i)




